from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from audit_common import AUDIT_DIR, REPORTS_DIR, ensure_dirs, load_json, write_csv, write_json

PROGRESS_PREFIX = "AUDIT_PROGRESS "


def emit_progress(current: int, total: int, message: str) -> None:
    """Emit structured progress for extension UI"""
    percent = round((current / total) * 100, 2) if total > 0 else 0
    payload = {
        "stage": "05_build_reports.py",
        "current": current,
        "total": total,
        "percent": percent,
        "message": message,
    }
    print(f"{PROGRESS_PREFIX}{json.dumps(payload, ensure_ascii=False)}", flush=True)


def write_xlsx(summary: dict, master_rows: list[dict], unmatched_rows: list[dict], 
               dam_dupes: list[dict], dam_phash_dupes: list[dict], 
               citizens_dupes: list[dict], governance: dict, output: Path) -> None:
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    for key, value in summary.items():
        ws_summary.append([key, value])
    
    # Add governance metrics
    ws_summary.append(["--- Governance Metrics ---", ""])
    for key, value in governance.items():
        ws_summary.append([key, value])
    
    ws_summary["A1"].font = Font(bold=True)
    ws_summary["B1"].font = Font(bold=True)

    # Citizens Images sheet
    ws_master = wb.create_sheet("Citizens Images")
    headers = [
        "image_url",
        "match_status",
        "match_method",
        "url_contains_asset_id",
        "dam_item_id",
        "dam_file_name",
        "phash_distance",
        "page_count",
        "page_urls",
        "needs_dam_upload",
    ]
    ws_master.append(headers)
    for cell in ws_master[1]:
        cell.font = Font(bold=True)
    for row in master_rows:
        ws_master.append([row.get(h) for h in headers])

    ws_unmatched = wb.create_sheet("Needs DAM Upload")
    ws_unmatched.append(headers)
    for cell in ws_unmatched[1]:
        cell.font = Font(bold=True)
    for row in unmatched_rows:
        ws_unmatched.append([row.get(h) for h in headers])

    ws_dupes = wb.create_sheet("DAM Duplicates (Exact)")
    dup_headers = ["sha256", "count", "item_ids", "file_names"]
    ws_dupes.append(dup_headers)
    for cell in ws_dupes[1]:
        cell.font = Font(bold=True)
    for row in dam_dupes:
        ws_dupes.append([
            row.get("sha256"),
            row.get("count"),
            " | ".join(row.get("item_ids", [])),
            " | ".join(row.get("file_names", [])),
        ])
    
    # DAM Phash Duplicates (visually similar)
    ws_phash_dupes = wb.create_sheet("DAM Duplicates (Similar)")
    phash_dup_headers = ["phash_group", "count", "item_ids", "file_names"]
    ws_phash_dupes.append(phash_dup_headers)
    for cell in ws_phash_dupes[1]:
        cell.font = Font(bold=True)
    for row in dam_phash_dupes:
        ws_phash_dupes.append([
            " | ".join(row.get("phash_group", [])),
            row.get("count"),
            " | ".join(row.get("item_ids", [])),
            " | ".join(row.get("file_names", [])),
        ])
    
    # Citizens Duplicates sheet (same image, multiple URLs)
    ws_citizens_dupes = wb.create_sheet("Citizens Duplicates")
    citizens_dup_headers = ["phash", "count", "image_urls", "dam_item_id", 
                           "total_page_count", "has_direct_dam_url", "has_local_copy"]
    ws_citizens_dupes.append(citizens_dup_headers)
    for cell in ws_citizens_dupes[1]:
        cell.font = Font(bold=True)
    for row in citizens_dupes:
        ws_citizens_dupes.append([
            row.get("phash"),
            row.get("count"),
            " | ".join(row.get("image_urls", [])),
            row.get("dam_item_id"),
            row.get("total_page_count"),
            "Yes" if row.get("has_direct_dam_url") else "No",
            "Yes" if row.get("has_local_copy") else "No",
        ])

    wb.save(output)


def write_html(master_rows: list[dict], summary: dict, governance: dict, output: Path) -> None:
    payload = json.dumps(master_rows, ensure_ascii=False)
    summary_payload = json.dumps(summary, ensure_ascii=False)
    
    # Format governance metrics
    total_matched = governance.get("total_matched_images", 0)
    direct_dam = governance.get("using_direct_dam_urls", 0)
    local_copies = governance.get("using_local_copies", 0)
    adoption_rate = governance.get("dam_url_adoption_rate", 0)
    dupe_groups = governance.get("citizens_duplicate_groups", 0)
    dupe_urls = governance.get("total_duplicate_urls", 0)
    
    # Extract summary metrics
    total_images = summary.get("citizens_images_total", 0)
    matched_direct = summary.get("matched_url_direct", 0)
    matched_exact = summary.get("matched_exact", 0)
    matched_phash = summary.get("matched_phash", 0)
    unmatched = summary.get("unmatched", 0)
    needs_upload = summary.get("needs_dam_upload", 0)
    dam_exact_dupes = summary.get("dam_internal_dupe_groups", 0)
    dam_phash_dupes = summary.get("dam_phash_dupe_groups", 0)
    citizens_dupes = summary.get("citizens_duplicate_groups", 0)
    
    html = f"""<!doctype html>
<html>
<head>
  <meta charset=\"utf-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
  <title>Citizens vs DAM Audit Report</title>
  <style>
    body {{ font-family: Segoe UI, Arial, sans-serif; margin: 16px; background: #f8f9fa; }}
    .container {{ max-width: 1400px; margin: 0 auto; background: white; padding: 24px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
    h1 {{ margin-top: 0; color: #212529; }}
    h2 {{ color: #495057; border-bottom: 2px solid #dee2e6; padding-bottom: 8px; margin-top: 32px; }}
    .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 12px; margin: 24px 0; }}
    .summary-item {{ background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 6px; padding: 12px 16px; cursor: pointer; transition: all 0.2s; display: flex; justify-content: space-between; align-items: center; }}
    .summary-item:hover {{ background: #e9ecef; border-color: #adb5bd; transform: translateY(-2px); box-shadow: 0 4px 8px rgba(0,0,0,0.1); }}
    .summary-item .label {{ font-size: 13px; color: #6c757d; font-weight: 500; }}
    .summary-item .value {{ font-size: 24px; font-weight: bold; color: #212529; }}
    .summary-item.primary {{ border-left: 4px solid #007bff; }}
    .summary-item.success {{ border-left: 4px solid #28a745; }}
    .summary-item.warning {{ border-left: 4px solid #ffc107; }}
    .summary-item.danger {{ border-left: 4px solid #dc3545; }}
    .summary-item.info {{ border-left: 4px solid #17a2b8; }}
    .controls {{ display: flex; gap: 8px; margin: 16px 0 12px 0; flex-wrap: wrap; align-items: center; }}
    .controls button {{ padding: 8px 16px; background: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 13px; }}
    .controls button:hover {{ background: #0056b3; }}
    input, select {{ padding: 6px 10px; border: 1px solid #ced4da; border-radius: 4px; font-size: 13px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 12px; }}
    th, td {{ border: 1px solid #ddd; padding: 8px; vertical-align: middle; }}
    th {{ background: #f3f3f3; position: sticky; top: 0; font-weight: 600; text-align: left; }}
    .thumbnail {{ width: 80px; height: 60px; object-fit: cover; border-radius: 4px; display: block; }}
    .thumbnail-link {{ display: block; text-decoration: none; }}
    .thumbnail-link:hover .thumbnail {{ opacity: 0.8; box-shadow: 0 2px 8px rgba(0,0,0,0.2); }}
    .pill {{ padding: 3px 8px; border-radius: 12px; font-size: 11px; color: #333; font-weight: 500; display: inline-block; }}
    .match_url_direct {{ background: #c3f0ca; color: #0d5e1e; }}
    .match_exact {{ background: #dff6e7; color: #155724; }}
    .match_phash {{ background: #fff6dd; color: #856404; }}
    .unmatched {{ background: #fde8e8; color: #721c24; }}
    .badge {{ padding: 2px 6px; background: #e0e0e0; border-radius: 3px; font-size: 10px; font-weight: 500; }}
    .badge-dam {{ background: #c3f0ca; color: #0d5e1e; }}
    .badge-local {{ background: #fff6dd; color: #856404; }}
    .page-urls {{ font-size: 11px; max-width: 300px; }}
    .page-urls a {{ color: #007bff; text-decoration: none; }}
    .page-urls a:hover {{ text-decoration: underline; }}
    #count {{ font-size: 14px; color: #6c757d; margin-left: auto; }}
    .dashboard {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin: 16px 0; }}
    .metric-card {{ background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 8px; padding: 16px; }}
    .metric-card h3 {{ margin: 0 0 8px 0; font-size: 14px; color: #6c757d; text-transform: uppercase; }}
    .metric-card .value {{ font-size: 32px; font-weight: bold; color: #212529; margin: 8px 0; }}
    .metric-card .label {{ font-size: 12px; color: #6c757d; }}
    .metric-card.success {{ border-left: 4px solid #28a745; }}
    .metric-card.warning {{ border-left: 4px solid #ffc107; }}
    .metric-card.info {{ border-left: 4px solid #17a2b8; }}
  </style>
</head>
<body>
  <div class="container">
    <h1>📊 Citizens vs DAM Audit Report</h1>
    
    <h2>Summary</h2>
    <div class="summary-grid">
      <div class="summary-item primary" data-filter="all">
        <span class="label">Total Citizens Images</span>
        <span class="value">{total_images}</span>
      </div>
      <div class="summary-item success" data-filter="match_url_direct">
        <span class="label">✅ Direct DAM URLs</span>
        <span class="value">{matched_direct}</span>
      </div>
      <div class="summary-item success" data-filter="match_exact">
        <span class="label">✅ Matched (Exact)</span>
        <span class="value">{matched_exact}</span>
      </div>
      <div class="summary-item warning" data-filter="match_phash">
        <span class="label">⚠️ Matched (Similar)</span>
        <span class="value">{matched_phash}</span>
      </div>
      <div class="summary-item danger" data-filter="unmatched">
        <span class="label">❌ Unmatched</span>
        <span class="value">{unmatched}</span>
      </div>
      <div class="summary-item danger" data-filter="needs_upload">
        <span class="label">📤 Needs DAM Upload</span>
        <span class="value">{needs_upload}</span>
      </div>
      <div class="summary-item info" data-filter="dam_exact_dupes">
        <span class="label">🔄 DAM Duplicates (Exact)</span>
        <span class="value">{dam_exact_dupes}</span>
      </div>
      <div class="summary-item info" data-filter="dam_phash_dupes">
        <span class="label">🔄 DAM Duplicates (Similar)</span>
        <span class="value">{dam_phash_dupes}</span>
      </div>
      <div class="summary-item info" data-filter="citizens_dupes">
        <span class="label">🔄 Citizens Duplicates</span>
        <span class="value">{citizens_dupes}</span>
      </div>
    </div>
    
    <h2>Governance Metrics</h2>
    <div class="dashboard">
      <div class="metric-card success">
        <h3>DAM Adoption</h3>
        <div class="value">{adoption_rate}%</div>
        <div class="label">{direct_dam} of {total_matched} using direct DAM URLs</div>
      </div>
      <div class="metric-card warning">
        <h3>Local Copies</h3>
        <div class="value">{local_copies}</div>
        <div class="label">Images downloaded instead of direct DAM links</div>
      </div>
      <div class="metric-card info">
        <h3>Duplicate Groups</h3>
        <div class="value">{dupe_groups}</div>
        <div class="label">{dupe_urls} duplicate URLs found</div>
      </div>
    </div>
    
    <h2>Image Details</h2>
    <div class=\"controls\">
      <button id=\"clearFilters\">Clear All Filters</button>
      <input id=\"q\" placeholder=\"Search image URL or DAM item\" style=\"min-width:320px\" />
      <select id=\"status\">
      <option value=\"\">All statuses</option>
      <option value=\"match_url_direct\">match_url_direct (Direct DAM)</option>
      <option value=\"match_exact\">match_exact</option>
      <option value=\"match_phash\">match_phash</option>
      <option value=\"unmatched\">unmatched</option>
      <option value=\"unmatched_error\">unmatched_error</option>
    </select>
    <select id=\"urlType\">
      <option value=\"\">All URL types</option>
      <option value=\"direct_dam\">Direct DAM URLs</option>
      <option value=\"local_copy\">Local Copies</option>
    </select>
    <label><input type=\"checkbox\" id=\"needs\" /> Needs DAM Upload only</label>
      <div id=\"count\"></div>
    </div>
    <table>
      <thead>
        <tr>
          <th>Preview</th>
          <th>Image URL</th>
          <th>Status</th>
          <th>URL Type</th>
          <th>DAM Item ID</th>
          <th>DAM File Name</th>
          <th>pHash Dist</th>
          <th>Pages</th>
          <th>Page URLs</th>
        </tr>
      </thead>
      <tbody id=\"rows\"></tbody>
    </table>
  <script src=\"report.js\"></script>
  <script>
    // Initialize report with data
    initializeReport({payload}, {summary_payload});
  </script>
</body>
</html>"""
    output.write_text(html, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build final CSV/XLSX/HTML audit reports")
    parser.add_argument("--matches", type=Path, default=AUDIT_DIR / "match_results.json")
    parser.add_argument("--unmatched", type=Path, default=AUDIT_DIR / "unmatched_results.json")
    parser.add_argument("--dam-dupes", type=Path, default=AUDIT_DIR / "dam_internal_dupes.json")
    parser.add_argument("--dam-phash-dupes", type=Path, default=AUDIT_DIR / "dam_phash_dupes.json")
    parser.add_argument("--citizens-dupes", type=Path, default=AUDIT_DIR / "citizens_duplicates.json")
    parser.add_argument("--governance", type=Path, default=AUDIT_DIR / "governance_metrics.json")
    args = parser.parse_args()

    ensure_dirs()
    
    # Total steps for progress tracking
    total_steps = 6
    
    emit_progress(0, total_steps, "Loading match results...")
    matches = load_json(args.matches)
    unmatched = load_json(args.unmatched)
    dam_dupes = load_json(args.dam_dupes)
    
    # Load new governance data (may not exist in older runs)
    try:
        dam_phash_dupes = load_json(args.dam_phash_dupes)
    except FileNotFoundError:
        dam_phash_dupes = []
    
    try:
        citizens_dupes = load_json(args.citizens_dupes)
    except FileNotFoundError:
        citizens_dupes = []
    
    try:
        governance = load_json(args.governance)
    except FileNotFoundError:
        governance = {}

    emit_progress(1, total_steps, "Preparing report data...")
    master_rows: list[dict] = []
    for row in matches:
        out = dict(row)
        out["needs_dam_upload"] = False
        out["page_urls"] = "|".join(out.get("page_urls", []))
        master_rows.append(out)
    for row in unmatched:
        out = dict(row)
        out["needs_dam_upload"] = out.get("match_status") in {"unmatched", "unmatched_error"}
        out["page_urls"] = "|".join(out.get("page_urls", []))
        master_rows.append(out)

    summary = {
        "citizens_images_total": len(master_rows),
        "matched_url_direct": sum(1 for r in master_rows if r.get("match_status") == "match_url_direct"),
        "matched_exact": sum(1 for r in master_rows if r.get("match_status") == "match_exact"),
        "matched_phash": sum(1 for r in master_rows if r.get("match_status") == "match_phash"),
        "unmatched": sum(1 for r in master_rows if r.get("match_status") in {"unmatched", "unmatched_error"}),
        "needs_dam_upload": sum(1 for r in master_rows if r.get("needs_dam_upload")),
        "dam_internal_dupe_groups": len(dam_dupes),
        "dam_phash_dupe_groups": len(dam_phash_dupes),
        "citizens_duplicate_groups": len(citizens_dupes),
    }

    emit_progress(2, total_steps, "Generating CSV reports...")
    master_csv = AUDIT_DIR / "audit_master.csv"
    write_csv(
        master_csv,
        master_rows,
        ["image_url", "match_status", "match_method", "url_contains_asset_id", "dam_item_id", "dam_file_name", "phash_distance", "page_count", "page_urls", "needs_dam_upload"],
    )
    write_json(AUDIT_DIR / "audit_master.json", master_rows)
    write_json(AUDIT_DIR / "audit_summary.json", summary)

    emit_progress(3, total_steps, "Generating Excel report...")
    xlsx_out = REPORTS_DIR / "citizens_dam_audit.xlsx"
    unmatched_rows = [r for r in master_rows if r.get("needs_dam_upload")]
    write_xlsx(summary, master_rows, unmatched_rows, dam_dupes, dam_phash_dupes, citizens_dupes, governance, xlsx_out)
    
    emit_progress(4, total_steps, "Generating HTML dashboard...")
    html_out = REPORTS_DIR / "audit_report.html"
    write_html(master_rows, summary, governance, html_out)
    
    emit_progress(5, total_steps, "Finalizing reports...")
    emit_progress(6, total_steps, "Report generation complete")

    print(json.dumps({
        "summary": summary,
        "governance": governance,
        "outputs": {
            "master_csv": str(master_csv),
            "master_json": str(AUDIT_DIR / "audit_master.json"),
            "summary_json": str(AUDIT_DIR / "audit_summary.json"),
            "xlsx": str(xlsx_out),
            "html": str(html_out),
        },
    }, indent=2))


if __name__ == "__main__":
    main()
